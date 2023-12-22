import mechanicalsoup
from dotenv import load_dotenv
import os
import openpyxl
import pandas as pd
import api.filemaker_api as filemaker
from datetime import datetime

def get_browser():
    load_dotenv()
    LTI_USER = os.getenv('LTI_USER')
    LTI_PASS = os.getenv('LTI_PASS')
    browser = mechanicalsoup.StatefulBrowser()

    login_url = "https://tms.languagetesting.com/Clientsite/login.aspx"  # Replace with the actual login page URL
    print(browser.open(login_url))
    browser.select_form('form[id="form1"]')  # Replace with the actual form name or attributes
    browser["txtLoginID"] = LTI_USER  # Replace with your username
    browser["txtPassword"] = LTI_PASS  # Replace with your password
    browser.submit_selected()
    return browser
def get_all_data(browser, fromdate, todate, byuid, language, kwargs):
    download_url = "https://tms.languagetesting.com/clientsite/DownloadRatingsV2.aspx"
    print(browser.open(download_url))
    browser.select_form('form[id="form1"]')
    browser['ctl00$cphMain$txtFromDate'] = fromdate
    browser['ctl00$cphMain$txtToDate'] = todate
    if byuid != None:
        browser['ctl00$cphMain$txtCandidateID'] = byuid
    if language != None:
        browser['ctl00$cphMain$ddlLanguages'] = language
    response = browser.submit_selected()
    print(response.status_code)
    response.raise_for_status()

    file_path = 'all-tests' + '.xlsx'
    with open(file_path, 'wb') as outf:
        outf.write(response.content)

    xlsx_file = openpyxl.load_workbook(file_path, data_only=True)

    sheet = xlsx_file.active

    firstname_index = 0
    lastname_index = 1
    byuid_index = 2
    BYU_ON_LINE_index = 3
    language_index = 4
    test_type_index = 5
    test_date_index = 6
    score_index = 7
    note_index = 8
    self_assessment_index = 9
    form_name_index = 10
    aappl_index = 11
    language_background_index = 12
    unit_index = 13
    major_index = 14
    minor_index = 15
    major2_index = 16
    minor2_index = 17
    special_instructions_index = 18
    testing_date_index = 19

    variable_mapping = {}
    for key, value in kwargs.items():
        if key in ['firstname', 'lastname', 'byuid', 'BYU_ON_LINE', 'language', 'test_type',
                'test_date', 'score', 'note', 'self_assessment', 'form_name', 'aappl',
                'language_background', 'unit', 'major', 'minor', 'major2', 'minor2',
                'special_instructions', 'testing_date']:
            variable_mapping[key] = locals()[f'{key}_index']

    data_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {}
        for variable_name, column_index in variable_mapping.items():
            cell_value = row[column_index]
            if cell_value is not None and isinstance(cell_value, str):
                cell_value = cell_value.strip()
            data[variable_name] = cell_value
        data_list.append(data)


    # if os.path.exists(file_path):
    #     os.remove(file_path)
    return data_list


def get_data(browser, fromdate, todate, test_type, byuid, language, kwargs):
    try:
        print(language)
        if test_type == 'WPT':
            test_type = 'WPT Test (no grid)'
        if test_type == 'OPI':
            test_type = 'OPI Test'
        if test_type == 'OPIc':
            test_type = 'OPIc Test'
            
        download_url = "https://tms.languagetesting.com/clientsite/DownloadRatingsV2.aspx"
        print(browser.open(download_url))
        browser.select_form('form[id="form1"]')
        browser['ctl00$cphMain$txtFromDate'] = fromdate
        browser['ctl00$cphMain$txtToDate'] = todate
        browser['ctl00$cphMain$ddlTestOptions'] = test_type
        if byuid != None:
            browser['ctl00$cphMain$txtCandidateID'] = byuid
        if language != None:
            browser['ctl00$cphMain$ddlLanguages'] = language

        response = browser.submit_selected()
        print(response.status_code)
        response.raise_for_status()

        file_path = test_type + '.xlsx'
        with open(file_path, 'wb') as outf:
            outf.write(response.content)

        xlsx_file = openpyxl.load_workbook(file_path, data_only=True)

        sheet = xlsx_file.active

        firstname_index = 0
        lastname_index = 1
        byuid_index = 2
        BYU_ON_LINE_index = 3
        language_index = 4
        test_type_index = 5
        test_date_index = 6
        score_index = 7
        note_index = 8
        self_assessment_index = 9
        form_name_index = 10
        aappl_index = 11
        language_background_index = 12
        unit_index = 13
        major_index = 14
        minor_index = 15
        major2_index = 16
        minor2_index = 17
        special_instructions_index = 18
        testing_date_index = 19

        variable_mapping = {}
        for key, value in kwargs.items():
            if key in ['firstname', 'lastname', 'byuid', 'BYU_ON_LINE', 'language', 'test_type',
                    'test_date', 'score', 'note', 'self_assessment', 'form_name', 'aappl',
                    'language_background', 'unit', 'major', 'minor', 'major2', 'minor2',
                    'special_instructions', 'testing_date']:
                variable_mapping[key] = locals()[f'{key}_index']

        data_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {}
            for variable_name, column_index in variable_mapping.items():
                cell_value = row[column_index]
                if cell_value is not None and isinstance(cell_value, str):
                    cell_value = cell_value.strip()
                data[variable_name] = cell_value
            data_list.append(data)
    except Exception as e:
        print(e)
        data_list = []

    # if os.path.exists(file_path):
    #     os.remove(file_path)
    return data_list
    
def close_browser(browser):
    browser.close()


def query_record(firstname, lastname, df, test_type, record_date_valid_by):
    query_date_object  = datetime.strptime(record_date_valid_by, '%m/%d/%Y')

    result = df.query("`Candidate First Name` == @firstname and `Candidate Last Name` == @lastname")
    if not result.empty:
        most_recent_entry = None

        for index, row in result.iterrows():
            if most_recent_entry is None or row["Test Date"] > most_recent_entry["Test Date"]:
                most_recent_entry = row

        found_name = most_recent_entry["Candidate First Name"] + " " + most_recent_entry["Candidate Last Name"]
        test_date = most_recent_entry["Test Date"]
        language = most_recent_entry["Language"]
        rating = most_recent_entry["Rating"]
        record_date_object = datetime.strptime(test_date, '%m/%d/%Y')
        if query_date_object <= record_date_object and "NS" not in rating and "NR" not in rating:
            return rating.strip(), test_date, language
        else:
            print("Record too old.")
            print(found_name, record_date_object, query_date_object)
            
    else:
        pass

def start_search(record_date_valid_by, semester_date_valid_by):
    opi = None
    wpt = None
    opic = None
    try:
        opi_df = pd.read_excel('OPI Test.xlsx')
        wpt_df = pd.read_excel('WPT Test (no grid).xlsx')
        opic_df = pd.read_excel('OPIc Test.xlsx')
    except Exception as e:
        print(e)
    token = filemaker.login()
    record_response = filemaker.get_all_by_date(token)
    filemaker.logout(token)
    student_list = []
    for student in record_response['response']['data']:
        if student['fieldData']['EntryDate'] >= semester_date_valid_by and student['fieldData']['CertificateStatus'] != "Email" and student['fieldData']['CertificateStatus'] != "Awarded" and student['fieldData']['CertificateStatus'] != "Unqualified":
            firstname = student['fieldData']['FirstName']
            lastname = student['fieldData']['LastName']
            abbre_language = student['fieldData']['Language']
            byuid = student['fieldData']['BYUID']
            reason = student['fieldData']['Reason']
            netid = student['fieldData']['NetID']
            recordId = student['recordId']
            try:
                opi = query_record(firstname, lastname, opi_df, 'OPI', record_date_valid_by)
                wpt = query_record(firstname, lastname, wpt_df, 'WPT', record_date_valid_by)
                opic = query_record(firstname, lastname, opic_df, 'OPIc', record_date_valid_by)
            except Exception as e:
                print(e)
            print(recordId, firstname, lastname, abbre_language, "OPI:", opi, "WPT:", wpt, "OPIc:", opic)
            if (opi != None or opic!= None) and wpt != None:
                student_data = {"recordId": recordId, "firstname": firstname, "lastname": lastname, "byuid": byuid, "netid":netid, "reason":reason, "language": {abbre_language: wpt[2]}, "opiScore": opi[0] if opi != None else 'None',
                    "opiDate": opi[1] if opi != None else 'None', "wptScore": wpt[0], "wptDate": wpt[1], "opicScore": opic[0] if opic != None else 'None', "opicDate": opic[1] if opic != None else 'None'}
                student_list.append(student_data)
        else:
            pass
    return student_list


# browser = get_browser()

# data = get_data(browser, fromdate='04/29/2023', todate='10/28/2023', test_type='WPT Test (no grid)', kwargs={'firstname':None, 'score': None, 'major': None})
# print(data)

# close_browser(browser)


